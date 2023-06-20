from openpyxl import Workbook, load_workbook ## 新增到excel所需要的模組 (20行~...用到)

wb = load_workbook('add_to_excel_test.xlsx') ##打開在這個資料夾中的excel -- wb=w(ork)b(ook)
ws = wb['全榖雜糧類'] ## ws=work sheet 就是乳品~豆、魚、蛋、肉類~等等 excel下面的不同工作表(work sheet)


##--------------------判斷有幾row---------------------------------   
row_len = 0                                ##判斷目前row有多長 (資料數目+1 最上面的類別也站一格 )                 ##r1c1 r1c2 r1c3
for row in range(1,40):                    ##假設不超過40比                                                    ##r2c1 r2c2 r2c3
    if ws['A'+str(row)].value !=None :     ##從A1(食物名稱)到An (不為空的)                                      ##r3c1 r3c2 r3c3
        row_len +=1                                                                                           ##r4c1 r4c2 r4c3 (row max = 4)
    else:                                          
        break
#print(row_len)


##-------------檢驗重複 沒有就新增--------------------      
repeat = False                                  
for row in range (1,row_len+1): ##--------------------從A1開始比對到An
    if(ws['A'+str(row)].value == "大米"):  ##------ws[A1].value 是說 cellA1的值
        repeat = True## ----------------------------A1 A2 A3 ...往下檢驗有沒有重複 
        break                               ##-------一旦有就可以跳出迴圈了
if(repeat): ##----------------------------如果ws[A3].value (名稱)是 和輸入重複的
        ws['B'+str(row)].value=30   ##--- 那 B3 得值(價格) 更改成新的
        ws['C'+str(row)].value="g"  ## ---C3(單位)的值也更新
        repeat = False 
else:
    ws._current_row = row_len          ##append到現在的格數的下一格
    ws.append(["高麗菜",'20','克'])     ##都沒有重複的話 就在最後新增

##-------------檢驗重複 沒有就新增------------------------


##----------刪除2~8列---------
#for i in wb:
#    i.delete_rows(5,6)
##----------刪除2~8列---------


wb.save('add_to_excel_test.xlsx')    ##存檔