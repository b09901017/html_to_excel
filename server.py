from flask import Flask, request ## 建立後端伺服器所需要的模組
from openpyxl import Workbook, load_workbook ## 新增到excel所需要的模組 (21行~53行)

app = Flask(__name__)

##--------主頁~~-------------
@app.route('/')
def homepage():
    return 'this is home page!'

##---------按下送出後會來到這裡-----------------
@app.route('/submit')
def submit():

    ##-----------把前端表單的東東存在變數裡--------------
    food_class = request.args.get("class") ## 用 request.args.get 來取得<input name = "name">標籤中 name = "的東東"
    name = request.args.get("name") ## 用 request.args.get 來取得標籤name = "的東東"
    price = request.args.get("price") ## 用 request.args.get 來取得標籤name = "的東東"
    unit = request.args.get("unit") ## 用 request.args.get 來取得標籤name = "的東東"

    ##-----------------------------------------------把拿到的變數新增到excel中------------------------------------------------------------------------------
    
    wb = load_workbook('add_to_excel_test.xlsx') ##打開在這個資料夾中的excel -- wb=w(ork)b(ook)
    ws = wb[food_class] ## ws=work sheet 就是乳品~豆、魚、蛋、肉類~等等 excel下面的不同工作表(work sheet)

    ##--------------------判斷有幾row---------------------------------   
    row_len = 0                                ##判斷目前row有多長 (資料數目+1 最上面的類別也站一格 )                 ##r1c1 r1c2 r1c3
    for row in range(1,40):                    ##假設不超過40比                                                    ##r2c1 r2c2 r2c3
        if ws['A'+str(row)].value !=None :     ##從A1(食物名稱)到An (不為空的)                                      ##r3c1 r3c2 r3c3
            row_len +=1                                                                                           ##r4c1 r4c2 r4c3 (row max = 4)
        else:                                          
            break
    #print(row_len)

    ##-------------檢驗重複 沒有就新增--------------------      
    repeat = False                              
    for row in range (1,row_len+1): ##--------------------從A1開始比對到An
        if(ws['A'+str(row)].value == name):  ##------ws[A1].value 是說 cellA1的值
            repeat = True## ----------------------------A1 A2 A3 ...往下檢驗有沒有重複 
            repeat_row = row
            break                               ##-------一旦有就可以跳出迴圈了
    if(repeat): ##----------------------------如果ws[A3].value (名稱)是 和輸入重複的
            ws['B'+str(repeat_row)].value=price   ##--- 那 B3 得值(價格) 更改成新的
            ws['C'+str(repeat_row)].value=unit  ## ---C3(單位)的值也更新
            repeat = False 
    else:
        ws._current_row = row_len          ##append到現在的格數的下一格
        ws.append([name,price,unit])     ##都沒有重複的話 就在最後新增

    ##-------------檢驗重複 沒有就新增------------------------

    ##----------刪除2~8列---------
    #for i in wb:
    #   i.delete_rows(2,11)
    ##----------刪除2~8列---------

    wb.save('add_to_excel_test.xlsx')    ##存檔

##--------------------------------------------------把拿到的變數新增到excel中-------------------------------------------------------




    return '類別 : ' +food_class + ' ' + '名稱 : ' +name  + ' ' + '價格 : '+price   + ' ' + '單位 : ' + unit  
   

if __name__ == '__main__':
    app.run()